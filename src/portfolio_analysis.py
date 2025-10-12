"""
Stock Portfolio Analysis Script
Fetches current stock prices and calculates portfolio performance
Outputs results to an Excel file
"""

import pandas as pd
import yfinance as yf
from datetime import datetime


def get_stock_data(ticker):
    """Fetch current stock price and basic info"""
    try:
        stock = yf.Ticker(ticker)
        current_price = stock.info.get("currentPrice") or stock.info.get(
            "regularMarketPrice"
        )
        if current_price is None:
            # Try getting from history if info doesn't have price
            hist = stock.history(period="1d")
            if not hist.empty:
                current_price = hist["Close"].iloc[-1]
        return current_price
    except Exception as e:
        print(f"Error fetching data for {ticker}: {e}")
        return None


def analyze_portfolio(portfolio_data):
    """
    Analyze portfolio performance

    Parameters:
    portfolio_data: list of dicts with keys: 'ticker', 'shares', 'purchase_price'
    """
    results = []

    for holding in portfolio_data:
        ticker = holding["ticker"]
        shares = holding["shares"]
        purchase_price = holding["purchase_price"]

        print(f"Fetching data for {ticker}...")
        current_price = get_stock_data(ticker)

        if current_price:
            initial_value = shares * purchase_price
            current_value = shares * current_price
            profit_loss = current_value - initial_value
            profit_loss_pct = ((current_price - purchase_price) / purchase_price) * 100

            results.append(
                {
                    "Ticker": ticker,
                    "Shares": shares,
                    "Purchase Price": purchase_price,
                    "Current Price": round(current_price, 2),
                    "Initial Value": round(initial_value, 2),
                    "Current Value": round(current_value, 2),
                    "Profit/Loss": round(profit_loss, 2),
                    "Return %": round(profit_loss_pct, 2),
                }
            )
        else:
            results.append(
                {
                    "Ticker": ticker,
                    "Shares": shares,
                    "Purchase Price": purchase_price,
                    "Current Price": "N/A",
                    "Initial Value": round(shares * purchase_price, 2),
                    "Current Value": "N/A",
                    "Profit/Loss": "N/A",
                    "Return %": "N/A",
                }
            )

    return results


def export_to_excel(results, filename=None):
    """Export portfolio analysis to Excel file"""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"portfolio_analysis_{timestamp}.xlsx"

    df = pd.DataFrame(results)

    # Sort by Profit/Loss (descending - best performers first)
    # Handle 'N/A' values by treating them as 0 for sorting
    df_sorted = df.copy()
    df_sorted["PL_Sort"] = df_sorted["Profit/Loss"].apply(
        lambda x: x if x != "N/A" else -999999
    )
    df_sorted = df_sorted.sort_values("PL_Sort", ascending=False)
    df_sorted = df_sorted.drop("PL_Sort", axis=1)
    df = df_sorted

    # Calculate totals
    total_initial = df[df["Initial Value"] != "N/A"]["Initial Value"].sum()
    total_current = df[df["Current Value"] != "N/A"]["Current Value"].sum()
    total_pl = df[df["Profit/Loss"] != "N/A"]["Profit/Loss"].sum()
    total_return_pct = (
        ((total_current - total_initial) / total_initial) * 100
        if total_initial > 0
        else 0
    )

    # Add summary row
    summary = pd.DataFrame(
        [
            {
                "Ticker": "TOTAL",
                "Shares": "",
                "Purchase Price": "",
                "Current Price": "",
                "Initial Value": round(total_initial, 2),
                "Current Value": round(total_current, 2),
                "Profit/Loss": round(total_pl, 2),
                "Return %": round(total_return_pct, 2),
            }
        ]
    )

    df = pd.concat([df, summary], ignore_index=True)

    # Create Excel writer with formatting
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Portfolio", index=False)

        # Get the worksheet
        worksheet = writer.sheets["Portfolio"]

        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Apply formatting to header
        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Format summary row
        last_row = worksheet.max_row
        summary_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )
        summary_font = Font(bold=True)

        for cell in worksheet[last_row]:
            cell.fill = summary_fill
            cell.font = summary_font

    print(f"\n✓ Portfolio analysis exported to: {filename}")
    return filename, df


def main():
    """Main function to run portfolio analysis"""

    # EXAMPLE PORTFOLIO - Replace with your actual holdings
    portfolio = [
        {"ticker": "ITGR", "shares": 20, "purchase_price": 68.85},
        {"ticker": "ZIM", "shares": 15, "purchase_price": 17.63},
        {"ticker": "SMCI", "shares": 11, "purchase_price": 46.28},
        {"ticker": "LULU", "shares": 9, "purchase_price": 175.83},
        {"ticker": "V", "shares": 4, "purchase_price": 343.19},
        {"ticker": "UNH", "shares": 4, "purchase_price": 319.76},
        {"ticker": "REGN", "shares": 3, "purchase_price": 570.33},
        {"ticker": "META", "shares": 2, "purchase_price": 664.25},
        {"ticker": "MA", "shares": 2, "purchase_price": 558.50},
        # {'ticker': 'VUSA', 'shares': 4, 'purchase_price': 84.69},
        # {'ticker': 'AAPL', 'shares': 10, 'purchase_price': 150.00},
        # {'ticker': 'MSFT', 'shares': 5, 'purchase_price': 300.00},
        # {'ticker': 'GOOGL', 'shares': 8, 'purchase_price': 120.00},
        # {'ticker': 'TSLA', 'shares': 3, 'purchase_price': 250.00},
        # {'ticker': 'AMZN', 'shares': 4, 'purchase_price': 140.00},
    ]

    print("=" * 60)
    print("STOCK PORTFOLIO ANALYSIS")
    print("=" * 60)
    print(f"\nAnalyzing {len(portfolio)} holdings...\n")

    # Analyze portfolio
    results = analyze_portfolio(portfolio)
    print(results)
    # Export to Excel
    # filename, df = export_to_excel(results)

    # Display summary
    print("\n" + "=" * 60)
    print("PORTFOLIO SUMMARY")
    print("=" * 60)
    # print(df.to_string(index=False))
    print("=" * 60)


if __name__ == "__main__":
    main()
